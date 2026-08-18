"""Phase 0: extract the league draft workbook into committed JSON fixtures.

Usage:
    .venv/bin/python -m backend.scripts.history_import \
        --workbook ~/Downloads/"2026 Juiced Fantasy Baseball Draft.xlsx"

Writes to backend/data/fixtures/league_history/:
  drafts_YYYY.json    — every pick, with owner, round and pick number
  keepers_YYYY.json   — manager, player, round cost, seasons kept
  rankings_YYYY.json  — the ESPN/Yahoo top-300 snapshots
  manifest.json       — per sheet: rows parsed, layout variant, every issue

The workbook itself stays out of git (7.5 MB binary, and it keeps changing).
The extracted JSON is small, diffable and reviewable, which is the point: a
reviewer can see what the parser believed about a 2013 sheet without opening
Excel.

Nothing here repairs data. Sheets that do not parse cleanly still emit their
rows, and the manifest records exactly what went wrong and where.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from backend.analysis.history.workbook import (
    parse_draft_sheet,
    parse_keeper_sheet,
    parse_ranking_sheet,
    ranking_snapshot_date,
)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_OUT = REPO_ROOT / "backend" / "data" / "fixtures" / "league_history"

_DRAFT_SHEET_RE = re.compile(r"^(\d{4}) Draft$")
_KEEPER_SHEET_RE = re.compile(r"^(\d{4}) Keepers$")

# Ranking sheets whose season is not in the name. ESPN300 sits between the
# 2020 sheets and its top three (Acuna, Yelich, Trout) is the 2020 preseason
# board; recorded as inferred so the manifest shows it was not stated.
_RANKING_SEASON_HINTS = {
    "ESPN300": (2020, True),
    "Yahoo300": (2020, True),
    "2017 ESPN Top 300": (2017, False),
}

# Sheets deliberately not extracted, with the reason. Listed so the manifest
# distinguishes "we chose not to" from "the parser missed it".
_SKIPPED = {
    "2026 Draft": "empty planning template; the real 2026 draft is in "
                  "backend/data/fixtures/retro_2026/draft_state_2026.json",
    "2020 & 2021 Supplemental Draft": "supplemental draft, not the main draft",
}


def _git_sha() -> str | None:
    try:
        out = subprocess.run(["git", "rev-parse", "HEAD"], cwd=REPO_ROOT,
                             capture_output=True, text=True, check=True)
        return out.stdout.strip()
    except (subprocess.CalledProcessError, FileNotFoundError):
        return None


def _write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w") as f:
        json.dump(payload, f, indent=2, sort_keys=True)
        f.write("\n")


def _is_ranking_sheet(name: str) -> bool:
    lowered = name.lower()
    return ("espn" in lowered or "yahoo" in lowered) and "300" in lowered.replace(
        " ", "") or lowered.startswith("espn by adp")


def _ranking_season(name: str) -> tuple[int | None, bool]:
    if name in _RANKING_SEASON_HINTS:
        return _RANKING_SEASON_HINTS[name]
    iso = ranking_snapshot_date(name)
    if iso:
        # A February/March snapshot is the board for that same season.
        return int(iso[:4]), False
    match = re.search(r"(\d{4})", name)
    if match:
        return int(match.group(1)), False
    return None, False


def load_sheets(workbook: Path) -> dict[str, list[tuple]]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=False)
    return {name: list(wb[name].iter_rows(values_only=True))
            for name in wb.sheetnames}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    workbook = args.workbook.expanduser()
    if not workbook.exists():
        raise SystemExit(f"workbook not found: {workbook}")

    sheets = load_sheets(workbook)
    header = {
        "extracted_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "git_sha": _git_sha(),
        "workbook": workbook.name,
        "workbook_sheets": len(sheets),
    }

    manifest_sheets: list[dict] = []
    drafts: dict[int, dict] = {}
    keepers: dict[int, dict] = {}
    rankings: dict[int, list[dict]] = {}

    for name, rows in sheets.items():
        if name in _SKIPPED:
            manifest_sheets.append({
                "sheet": name, "status": "skipped", "reason": _SKIPPED[name],
            })
            continue

        draft_match = _DRAFT_SHEET_RE.match(name)
        keeper_match = _KEEPER_SHEET_RE.match(name)

        if draft_match:
            season = int(draft_match.group(1))
            parsed = parse_draft_sheet(rows, season, name)
            drafts[season] = {
                **header,
                "season": season,
                "sheet": name,
                "layout_variant": parsed.variant,
                "columns": parsed.columns.as_dict(),
                "picks": parsed.picks,
            }
            manifest_sheets.append({
                "sheet": name,
                "status": "parsed",
                "kind": "draft",
                "season": season,
                "layout_variant": parsed.variant,
                "columns": parsed.columns.as_dict(),
                "rows_parsed": len(parsed.picks),
                "owners": parsed.owners,
                "owner_count": len(parsed.owners),
                "rounds": sorted({p["round"] for p in parsed.picks
                                  if p["round"] is not None}),
                "issues": [i.as_dict() for i in parsed.issues],
            })

        elif keeper_match:
            season = int(keeper_match.group(1))
            parsed = parse_keeper_sheet(rows, season, name)
            keepers[season] = {
                **header,
                "season": season,
                "sheet": name,
                "layout_variant": parsed.variant,
                "keepers": parsed.keepers,
                # Which draft slot each keeper consumed. Present only for 2015
                # and 2016, where it was pasted below the keeper table.
                "pick_slots": parsed.pick_slots,
            }
            manifest_sheets.append({
                "sheet": name,
                "status": "parsed",
                "kind": "keepers",
                "season": season,
                "layout_variant": parsed.variant,
                "rows_parsed": len(parsed.keepers),
                "managers": sorted({k["manager"] for k in parsed.keepers}),
                "pick_slots_parsed": len(parsed.pick_slots),
                "issues": [i.as_dict() for i in parsed.issues],
            })

        elif _is_ranking_sheet(name):
            season, inferred = _ranking_season(name)
            parsed = parse_ranking_sheet(rows, name, season, inferred)
            record = {
                "sheet": name,
                "snapshot_date": parsed.snapshot_date,
                "season_inferred": parsed.season_inferred,
                "rankings": parsed.rankings,
            }
            if season is not None:
                rankings.setdefault(season, []).append(record)
            manifest_sheets.append({
                "sheet": name,
                "status": "parsed" if season is not None else "unassigned",
                "kind": "rankings",
                "season": season,
                "season_inferred": inferred,
                "snapshot_date": parsed.snapshot_date,
                "rows_parsed": len(parsed.rankings),
                "issues": [i.as_dict() for i in parsed.issues],
            })

        else:
            manifest_sheets.append({
                "sheet": name, "status": "not_extracted",
                "reason": "no parser for this sheet family "
                          "(Trades, Rosters, Record Book)",
                "rows": sum(1 for r in rows if any(c is not None for c in r)),
            })

    out = args.out
    for season, payload in sorted(drafts.items()):
        _write_json(out / f"drafts_{season}.json", payload)
    for season, payload in sorted(keepers.items()):
        _write_json(out / f"keepers_{season}.json", payload)
    for season, records in sorted(rankings.items()):
        _write_json(out / f"rankings_{season}.json", {
            **header, "season": season, "sources": records,
        })

    total_issues = sum(len(s.get("issues", [])) for s in manifest_sheets)
    _write_json(out / "manifest.json", {
        **header,
        "draft_seasons": sorted(drafts),
        "keeper_seasons": sorted(keepers),
        "ranking_seasons": sorted(rankings),
        "total_issues": total_issues,
        "sheets": manifest_sheets,
    })

    # ── console summary ──
    print(f"\nParsed {len(sheets)} sheets from {workbook.name}\n")
    print(f"{'season':>7} {'picks':>6} {'owners':>7} {'keepers':>8} "
          f"{'variant':>12}  issues")
    for season in sorted(set(drafts) | set(keepers)):
        draft = drafts.get(season)
        keeper = keepers.get(season)
        issues = sum(len(s.get("issues", [])) for s in manifest_sheets
                     if s.get("season") == season)
        owners = len({p["owner"] for p in draft["picks"]}) if draft else 0
        print(f"{season:>7} {len(draft['picks']) if draft else 0:>6} "
              f"{owners:>7} {len(keeper['keepers']) if keeper else 0:>8} "
              f"{(draft or keeper or {}).get('layout_variant', '-'):>12}  {issues}")
    print(f"\nRanking snapshots: "
          f"{', '.join(str(s) for s in sorted(rankings)) or 'none'}")
    print(f"Total issues recorded: {total_issues}")
    print(f"Artifacts written to {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
